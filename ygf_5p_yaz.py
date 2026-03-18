# -*- coding: utf-8 -*-
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

import openpyxl
import gspread
from google.oauth2.service_account import Credentials
import time

CREDS = r"C:\Users\PDS\Desktop\snap code\credentials.json"
SHEET_ID = "1RvQnrukTy8LVxWjskFYUzlpCOuAqEU9IH7YTFM2XJQI"
SCOPES = ["https://www.googleapis.com/auth/spreadsheets", "https://www.googleapis.com/auth/drive"]
EXCEL_PATH = r"C:\Users\PDS\Desktop\claude\ygf\ygf liste.xlsx"

# ── 1) Excel'den verileri oku ──
wb = openpyxl.load_workbook(EXCEL_PATH)
ws = wb.active

# Satir 2: toplamlar, Satir 3: isimler, Satir 4+: hisseler
# Sütun 1: Endeks/HISSE, Sütun 2-12: kişiler (col index 2..12)
name_map = {}  # col_index -> isim
total_map = {}  # col_index -> toplam
for col in range(2, 13):
    name_raw = ws.cell(row=3, column=col).value
    total_raw = ws.cell(row=2, column=col).value
    if name_raw:
        name_map[col] = str(name_raw).strip()
        total_map[col] = float(str(total_raw).strip())

print("Yarismaci eslestirmeleri:")
for col, name in sorted(name_map.items()):
    print(f"  Col {col}: {name} (toplam: {total_map[col]})")

# Her kisi icin hisse-tutar listesi olustur
person_stocks = {}  # isim -> [(hisse, tutar), ...]
for col, name in name_map.items():
    stocks = []
    for row in range(4, ws.max_row + 1):
        hisse = ws.cell(row=row, column=1).value
        val = ws.cell(row=row, column=col).value
        if val is not None:
            # Tarih olarak okunan hucreleri toplamdan hesapla
            if hasattr(val, 'year'):
                other_sum = sum(
                    float(str(ws.cell(r, col).value))
                    for r in range(4, ws.max_row + 1)
                    if ws.cell(r, col).value is not None
                    and not hasattr(ws.cell(r, col).value, 'year')
                    and r != row
                )
                val = round(total_map[col] - other_sum, 2)
                print(f"  [DUZELTME] {name} - {hisse}: tarih -> {val}")
            else:
                val = float(str(val).strip())
            if val > 0:
                stocks.append((str(hisse).strip(), val))
    person_stocks[name] = stocks

print("\n" + "=" * 60)
print("EXCEL OZET:")
print("=" * 60)
for name, stocks in person_stocks.items():
    total = sum(t for _, t in stocks)
    pct_sum = sum(t / total_map[next(c for c, n in name_map.items() if n == name)] * 100 for _, t in stocks)
    print(f"\n{name} (toplam: {total:.2f}):")
    for hisse, tutar in stocks:
        col_idx = next(c for c, n in name_map.items() if n == name)
        pct = tutar / total_map[col_idx] * 100
        print(f"  {hisse:8s}  {tutar:>7.2f} TL  ({pct:>5.1f}%)")

# ── 2) Google Sheets'e yaz ──
creds = Credentials.from_service_account_file(CREDS, scopes=SCOPES)
gc = gspread.authorize(creds)
ss = gc.open_by_key(SHEET_ID)

# Worksheet isim eslestirmesi
ws_list = ss.worksheets()
ws_titles = {ws.title: ws for ws in ws_list}

# Excel'deki isimler ile Sheet'teki isimler eslestirilmeli
# Excel: Gürkan, Barış, Berkan, Turan, Serkan, Ali Cenk, Mehmet, Osman, Selim, Özhan, Oğuz
# Sheet'te ayni isimler var

print("\n" + "=" * 60)
print("GOOGLE SHEETS'E YAZILIYOR...")
print("=" * 60)

for name, stocks in person_stocks.items():
    # Sheet'te bu ismi bul
    target_ws = None
    for title, ws_obj in ws_titles.items():
        if name in title or title in name:
            target_ws = ws_obj
            break

    if target_ws is None:
        print(f"\n[HATA] '{name}' icin worksheet bulunamadi!")
        continue

    # 5. Periyot satirini bul
    all_vals = target_ws.get_all_values()
    periyot5_row = None
    toplam_row = None
    for i, row in enumerate(all_vals):
        if '5. Periyot' in str(row[0]):
            periyot5_row = i + 1  # 1-indexed
        if periyot5_row and i > periyot5_row and row[0] == 'TOPLAM':
            toplam_row = i + 1
            break

    if periyot5_row is None or toplam_row is None:
        print(f"\n[HATA] '{name}' sayfasinda 5. Periyot yapisi bulunamadi!")
        continue

    data_start = periyot5_row + 2  # header satirindan sonra
    available_rows = toplam_row - data_start
    needed_rows = len(stocks)

    print(f"\n{name} ({target_ws.title}):")
    print(f"  5.Periyot satir {periyot5_row}, veri baslangiç: {data_start}, TOPLAM: {toplam_row}")
    print(f"  Mevcut bos satir: {available_rows}, gereken: {needed_rows}")

    # Ek satir lazimsa ekle (TOPLAM satirindan once)
    if needed_rows > available_rows:
        extra = needed_rows - available_rows
        print(f"  {extra} ek satir ekleniyor...")
        target_ws.insert_rows([[''] * 8] * extra, row=toplam_row)
        toplam_row += extra
        time.sleep(1)

    # Toplam col index icin
    col_idx = next(c for c, n in name_map.items() if n == name)
    total = total_map[col_idx]

    # Hisseleri yaz
    cells_to_update = []
    for j, (hisse, tutar) in enumerate(stocks):
        row_num = data_start + j
        pct = round(tutar / total * 100, 1)
        pct_str = f"%{pct}"
        # A: Hisse, B: Agirlik %
        cells_to_update.append(gspread.Cell(row_num, 1, hisse))
        cells_to_update.append(gspread.Cell(row_num, 2, pct_str))
        print(f"  Satir {row_num}: {hisse} | {pct_str}")

    if cells_to_update:
        target_ws.update_cells(cells_to_update)
        print(f"  -> {len(stocks)} hisse yazildi!")
    time.sleep(2)  # rate limit

print("\n" + "=" * 60)
print("TAMAMLANDI!")
print("=" * 60)
