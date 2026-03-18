# -*- coding: utf-8 -*-
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

import openpyxl
import gspread
from google.oauth2.service_account import Credentials
import time

CREDS = r"C:\Users\PDS\Desktop\snap code\credentials.json"
SHEET_ID = "1RvQnrukTy8LVxWjskFYUzlpCOuAqEU9IH7YTFM2XJQI"
SCOPES = ["https://www.googleapis.com/auth/spreadsheets", "https://www.googleapis.com/auth/drive"]
EXCEL_PATH = r"C:\Users\PDS\Desktop\claude\ygf\ygf liste.xlsx"

# ── 1) Excel'den verileri oku ──
wb = openpyxl.load_workbook(EXCEL_PATH)
ws_xl = wb.active

name_map = {}
total_map = {}
for col in range(2, 13):
    name_raw = ws_xl.cell(row=3, column=col).value
    total_raw = ws_xl.cell(row=2, column=col).value
    if name_raw:
        name_map[col] = str(name_raw).strip()
        total_map[col] = float(str(total_raw).strip())

person_stocks = {}
for col, name in name_map.items():
    stocks = []
    for row in range(4, ws_xl.max_row + 1):
        hisse = ws_xl.cell(row=row, column=1).value
        val = ws_xl.cell(row=row, column=col).value
        if val is not None:
            if hasattr(val, 'year'):
                other_sum = sum(
                    float(str(ws_xl.cell(r, col).value))
                    for r in range(4, ws_xl.max_row + 1)
                    if ws_xl.cell(r, col).value is not None
                    and not hasattr(ws_xl.cell(r, col).value, 'year')
                    and r != row
                )
                val = round(total_map[col] - other_sum, 2)
            else:
                val = float(str(val).strip())
            if val > 0:
                stocks.append((str(hisse).strip(), val))
    person_stocks[name] = stocks

# ── 2) Google Sheets bağlan ──
creds = Credentials.from_service_account_file(CREDS, scopes=SCOPES)
gc = gspread.authorize(creds)
ss = gc.open_by_key(SHEET_ID)

ws_list = ss.worksheets()
ws_titles = {ws.title: ws for ws in ws_list}

# Yarışmacı isim eşleştirme
YARISMACILAR = ["Barış", "Serkan", "Ali Cenk", "Özhan", "Turan", "Berkan", "Selim", "Gürkan", "Osman", "Oğuz", "Mehmet"]

def find_ws(name):
    for title, ws_obj in ws_titles.items():
        if name == title or name in title:
            return ws_obj
    return None

# ── 3) Her yarışmacı sayfasını güncelle ──
for name, stocks in person_stocks.items():
    ws = find_ws(name)
    if ws is None:
        print(f"[HATA] '{name}' worksheet bulunamadi!")
        continue

    print(f"\n{'='*50}")
    print(f"  {ws.title}")
    print(f"{'='*50}")

    all_vals = ws.get_all_values()
    total_rows = len(all_vals)
    sheet_id = ws.id

    # ── 3a) Tüm periyotlarda "Lot" → "TL" header değişikliği ──
    lot_cells = []
    for i, row in enumerate(all_vals):
        if len(row) > 2 and row[2] == 'Lot':
            lot_cells.append(gspread.Cell(i + 1, 3, 'TL'))
            print(f"  Satir {i+1}: 'Lot' → 'TL'")

    if lot_cells:
        ws.update_cells(lot_cells)
        print(f"  {len(lot_cells)} header guncellendi")

    # ── 3b) 5. Periyot bölümünü bul ──
    periyot5_row = None
    toplam5_row = None
    for i, row in enumerate(all_vals):
        if '5. Periyot' in str(row[0]):
            periyot5_row = i + 1
        if periyot5_row and i + 1 > periyot5_row + 1 and row[0] == 'TOPLAM':
            toplam5_row = i + 1
            break

    if not periyot5_row or not toplam5_row:
        print(f"  [HATA] 5. Periyot yapisi bulunamadi!")
        continue

    data_start = periyot5_row + 2  # header'dan sonra
    data_end = toplam5_row - 1     # TOPLAM'dan önce

    print(f"  5P veri araligi: satir {data_start}-{data_end} ({data_end - data_start + 1} satir)")
    print(f"  Hisse sayisi: {len(stocks)}")

    # ── 3c) B sütununu temizle, C sütununa TL yaz, B'ye formül koy ──
    col_idx_xl = next(c for c, n in name_map.items() if n == name)
    total_tl = total_map[col_idx_xl]

    cells_to_update = []
    formul_requests = []

    for j, (hisse, tutar) in enumerate(stocks):
        row_num = data_start + j
        # A: Hisse (zaten yazili, tekrar yaz garanti olsun)
        cells_to_update.append(gspread.Cell(row_num, 1, hisse))
        # C: TL tutarı
        cells_to_update.append(gspread.Cell(row_num, 3, tutar))
        # B: Ağırlık % formülü (C_row / TOPLAM_C * 100 formatında)
        # =IF(C43="","",C43/C_TOPLAM*100) şeklinde
        # Ama TOPLAM satırında da SUM formülü olmalı
        # Formüller ayrı yazılacak
        formula = f'=IF(C{row_num}="","",ROUND(C{row_num}/C{toplam5_row}*100,1)&"%")'
        formul_requests.append({
            'row': row_num,
            'formula': formula
        })
        print(f"  Satir {row_num}: {hisse} | {tutar} TL")

    # Boş kalan satırları temizle (eski veri varsa)
    for row_num in range(data_start + len(stocks), data_end + 1):
        cells_to_update.append(gspread.Cell(row_num, 1, ''))
        cells_to_update.append(gspread.Cell(row_num, 2, ''))
        cells_to_update.append(gspread.Cell(row_num, 3, ''))

    # TOPLAM satırında C sütununa SUM formülü
    toplam_formula = f'=SUM(C{data_start}:C{data_end})'

    # Hücreleri güncelle (değerler)
    if cells_to_update:
        ws.update_cells(cells_to_update)

    # Formülleri yaz (batch update ile)
    formula_updates = []
    for req in formul_requests:
        cell_ref = f'B{req["row"]}'
        formula_updates.append({
            'range': cell_ref,
            'values': [[req['formula']]]
        })

    # TOPLAM C sütunu
    formula_updates.append({
        'range': f'C{toplam5_row}',
        'values': [[toplam_formula]]
    })

    # TOPLAM B sütunu (her zaman 100%)
    formula_updates.append({
        'range': f'B{toplam5_row}',
        'values': [['100%']]
    })

    if formula_updates:
        ws.batch_update(formula_updates, value_input_option='USER_ENTERED')

    print(f"  TOPLAM satir {toplam5_row}: SUM formulu eklendi")
    print(f"  -> Tamamlandi! ({len(stocks)} hisse, {total_tl} TL)")

    # ── 3d) Calibri 11pt siyah formatlama (tüm sayfa) ──
    format_request = {
        "requests": [
            {
                "repeatCell": {
                    "range": {
                        "sheetId": sheet_id,
                        "startRowIndex": 0,
                        "endRowIndex": total_rows + 5,
                        "startColumnIndex": 0,
                        "endColumnIndex": 8
                    },
                    "cell": {
                        "userEnteredFormat": {
                            "textFormat": {
                                "fontFamily": "Calibri",
                                "fontSize": 11,
                                "foregroundColorStyle": {
                                    "rgbColor": {
                                        "red": 0,
                                        "green": 0,
                                        "blue": 0
                                    }
                                }
                            }
                        }
                    },
                    "fields": "userEnteredFormat.textFormat(fontFamily,fontSize,foregroundColorStyle)"
                }
            }
        ]
    }
    ss.batch_update(format_request)
    print(f"  Calibri 11pt siyah format uygulandi")

    time.sleep(2)

print(f"\n{'='*50}")
print("  TAMAMLANDI!")
print(f"{'='*50}")
